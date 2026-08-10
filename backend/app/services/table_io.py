from __future__ import annotations

import csv
import io
import re
import zipfile

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from app.services.verification_links import attach as attach_verification_cases
from app.services.history import record_change


_HTML_TAG = re.compile(r"<[^>]*>")


def _strip_html(text: str) -> str:
    return _HTML_TAG.sub("", text).strip()


# Excel/LibreOffice evaluate a cell whose text begins with one of these, so a
# requirement named `=cmd|'/c calc'!A0` becomes a live formula in the export.
# Quoting doesn't prevent it — the value has to be neutralised.
_FORMULA_LEAD = ("=", "+", "-", "@", "\t", "\r")


def _defuse(value):
    """Prefix a leading formula trigger with an apostrophe (the spreadsheet
    convention for "this is literal text"). Non-strings pass through."""
    if isinstance(value, str):
        stripped = value.lstrip()
        if stripped.startswith(_FORMULA_LEAD):
            return "'" + value
    return value


def _defuse_row(row: dict) -> dict:
    return {k: _defuse(v) for k, v in row.items()}


def _flat_columns(meta: dict | None = None) -> list[str]:
    cols = [
        "id", "type", "name", "description", "status", "priority",
        "verification_method", "parent", "relations", "verification_cases",
        "rationale", "source", "allocated_to", "baselines",
    ]
    if meta and "attributes" in meta and "publish" in meta["attributes"]:
        for attr in meta["attributes"]["publish"]:
            if attr not in cols:
                cols.append(attr)
    return cols


def _req_to_row(req: dict, meta: dict | None = None) -> dict:
    row = {
        "id": req.get("id", ""),
        "type": req.get("type", "functional"),
        "name": req.get("name", ""),
        "description": _strip_html(req.get("description", "")),
        "status": req.get("status", "proposed"),
        "priority": req.get("priority", "medium"),
        "verification_method": req.get("verification_method", "test"),
        "parent": req.get("parent") or "",
        "relations": "; ".join(
            f"{r['type']}:{r['target']}" for r in req.get("relations", [])
        ),
        "verification_cases": "; ".join(req.get("verification_cases", [])),
        "rationale": req.get("rationale", ""),
        "source": req.get("source", ""),
        "allocated_to": req.get("allocated_to", ""),
        "baselines": ", ".join(req.get("baselines", [])),
    }
    for attr in req.get("attributes", []):
        row[attr["key"]] = attr["value"]
    return row


_HEADER_ALIASES: dict[str, str] = {
    "requirement_id": "id",
    "req_id": "id",
    "req id": "id",
    "requirement type": "type",
    "priority_level": "priority",
    "verification": "verification_method",
    "verification method": "verification_method",
    "v_cases": "verification_cases",
    "test_cases": "verification_cases",
    "linked_to": "relations",
    "linked requirements": "relations",
    "parent_id": "parent",
    "justification": "rationale",
    "allocated_to": "allocated_to",
    "allocated": "allocated_to",
    "system_element": "allocated_to",
    "baseline": "baselines",
}


def _row_to_req(row: dict) -> dict:
    names = {n.lower().strip().replace(" ", "_"): n for n in row}

    def get(key, default):
        # csv.DictReader pads a short row with None for the missing trailing
        # columns. `row.get(col, default)` returns that None rather than the
        # default (the key *is* present), so callers doing `.strip()` blew up
        # on any ragged row — after `mode=replace` had already deleted
        # everything. Coerce None to the caller's default.
        col = names.get(key) or names.get(_HEADER_ALIASES.get(key, ""))
        value = row.get(col, default) if col else default
        return default if value is None else value

    relations = []
    rel_str = get("relations", "")
    if rel_str.strip():
        for part in rel_str.split(";"):
            part = part.strip()
            if ":" in part:
                rtype, target = part.split(":", 1)
                relations.append({"type": rtype.strip(), "target": target.strip()})

    vcs = [v.strip() for v in get("verification_cases", "").split(";") if v.strip()]

    attributes = []
    for col_key, col_name in names.items():
        if col_key in {
            "id", "type", "name", "description", "status", "priority",
            "verification_method", "parent", "relations", "verification_cases",
            "rationale", "source", "allocated_to", "baselines", "created", "modified",
        }:
            continue
        val = row.get(col_name, "")
        if val:
            attributes.append({"key": col_name, "value": str(val)})

    req = {
        "id": get("id", ""),
        "type": get("type", "functional"),
        "name": get("name", ""),
        "description": get("description", ""),
        "status": get("status", "proposed"),
        "priority": get("priority", "medium"),
        "parent": get("parent", "") or None,
        "relations": relations,
        "verification_cases": vcs,
        "rationale": get("rationale", ""),
        "source": get("source", ""),
        "allocated_to": get("allocated_to", ""),
        "baselines": [b.strip() for b in get("baselines", "").split(",") if b.strip()] or [],
        "attributes": attributes,
    }
    return req


def export_table(store, fmt: str) -> str:
    if fmt not in ("csv", "tsv"):
        raise ValueError(f"Unknown table format: {fmt}")
    meta = store.read_meta()
    reqs = store.list_requirements()
    vcs = store.list_verification_cases()
    attach_verification_cases(store, reqs, vcs)
    columns = _flat_columns(meta)

    out = io.StringIO()
    delimiter = "\t" if fmt == "tsv" else ","
    writer = csv.DictWriter(out, fieldnames=columns, delimiter=delimiter, quoting=csv.QUOTE_ALL)
    writer.writeheader()
    for r in reqs:
        writer.writerow(_defuse_row(_req_to_row(r, meta)))
    return out.getvalue()


def _require_id_column(ids: list[str], rows: int) -> None:
    """Refuse to wipe the project for a file we got no usable ids out of.

    An unrecognised id column would otherwise delete everything and report a
    cheerful ``{"created": 0}``. This runs on dry runs too: a preview that says
    "57 will be deleted" for a file the real import rejects is worse than no
    preview, because it is the preview that talks the user into committing.
    """
    if rows and not any(ids):
        raise ValueError(
            "No 'id' column recognised in the table — refusing to replace "
            "the project. Check the header row."
        )


def _dry_run_counts(store, ids: list[str], mode: str) -> tuple[int, int, int, int]:
    """``(would_create, would_update, would_delete, skipped)`` for a preview.

    Walks the rows in order rather than counting set membership, so a file that
    repeats an id reports one create and one update — which is what the real
    import does on the second row, since the first has already landed by then.
    """
    would_delete = len(store.list_requirements()) if mode == "replace" else 0
    seen: set[str] = set()
    would_create = would_update = skipped = 0
    for rid in ids:
        if not rid:
            skipped += 1
            continue
        # In replace mode the store is emptied first, so nothing already in it
        # counts as an update — only a repeat within the file does.
        exists = rid in seen or (mode != "replace" and store.get_requirement(rid) is not None)
        if exists:
            would_update += 1
        else:
            would_create += 1
        seen.add(rid)
    return would_create, would_update, would_delete, skipped


def _dry_run_summary(store, ids: list[str], fmt: str, mode: str, rows: int) -> dict:
    """The one dry-run shape every table format returns.

    csv/tsv and xlsx used to disagree — xlsx put its counts in ``created``/
    ``updated`` while csv used ``would_create``/``would_update`` and zeroed the
    others — so a caller reading one was silently wrong about the other.
    """
    would_create, would_update, would_delete, skipped = _dry_run_counts(store, ids, mode)
    return {
        "created": 0,
        "updated": 0,
        "skipped": skipped,
        "traces_added": 0,
        "verification_cases": 0,
        "format": fmt,
        "dry_run": True,
        "would_create": would_create,
        "would_update": would_update,
        "would_delete": would_delete,
        "rows": rows,
    }


def import_table(store, content: str, fmt: str = "csv", mode: str = "merge",
                 dry_run: bool = False, *, username: str = "") -> dict:
    if fmt not in ("csv", "tsv"):
        raise ValueError(f"Unknown table format: {fmt}")
    if mode not in ("merge", "replace"):
        raise ValueError(f"Unknown mode: {mode}")

    delimiter = "\t" if fmt == "tsv" else ","
    reader = csv.DictReader(io.StringIO(content), delimiter=delimiter)
    rows = list(reader)

    created = 0
    updated = 0
    skipped = 0

    # Parse every row up-front. A malformed row used to raise *after* replace
    # mode had already wiped the collection, emptying the project and importing
    # nothing. Nothing is deleted until the whole file is known to be readable.
    try:
        parsed_rows = [_row_to_req(row) for row in rows]
    except Exception as exc:
        raise ValueError(f"Could not parse the table (nothing was changed): {exc}") from exc

    row_ids = [r.get("id", "").strip() for r in parsed_rows]

    if mode == "replace":
        _require_id_column(row_ids, len(rows))

    if dry_run:
        return _dry_run_summary(store, row_ids, fmt, mode, len(rows))

    if mode == "replace":
        for req in store.list_requirements():
            # Record after the delete succeeds, matching bulk_routes: logging
            # first would leave a delete entry for a record still present if the
            # write failed.
            if store.delete_requirement(req["id"]):
                record_change(store, req["id"], "delete", req, None, username)

    for req_data in parsed_rows:
        rid = req_data.get("id", "").strip()
        if not rid:
            skipped += 1
            continue

        existing = store.get_requirement(rid)
        if existing:
            store.update_requirement(rid, req_data)
            updated += 1
            record_change(store, rid, "update", existing, store.get_requirement(rid), username)
        else:
            store.create_requirement(req_data)
            created += 1
            record_change(store, rid, "create", None, req_data, username)

    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "traces_added": 0,
        "verification_cases": 0,
        "format": fmt,
    }


def export_xlsx(store, path: str) -> None:
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required for XLSX export. Install with: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "Requirements"

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    meta = store.read_meta()
    columns = _flat_columns(meta)
    reqs = store.list_requirements()
    vcs = store.list_verification_cases()
    attach_verification_cases(store, reqs, vcs)

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, r in enumerate(reqs, start=2):
        row_data = _req_to_row(r, meta)
        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=_defuse(row_data.get(col_name, "")))

    wb.save(path)


MAX_XLSX_UNCOMPRESSED_MB = 200
MAX_XLSX_COMPRESSION_RATIO = 100
MAX_XLSX_ESTIMATED_ROWS = 100_000


def _check_xlsx_zip_bomb(content: bytes) -> None:
    buf = io.BytesIO(content)
    try:
        with zipfile.ZipFile(buf) as zf:
            for info in zf.infolist():
                if info.filename.endswith("/"):
                    continue
                compressed = info.compress_size
                uncompressed = info.file_size
                if compressed <= 0:
                    continue
                ratio = uncompressed / compressed
                if ratio > MAX_XLSX_COMPRESSION_RATIO:
                    raise ValueError(
                        f"XLSX file rejected: compression ratio {ratio:.0f}:1 "
                        f"exceeds limit of {MAX_XLSX_COMPRESSION_RATIO}:1 — "
                        f"likely a zip bomb attempt"
                    )
                if uncompressed > MAX_XLSX_UNCOMPRESSED_MB * 1024 * 1024:
                    raise ValueError(
                        f"XLSX file rejected: uncompressed size "
                        f"{uncompressed / (1024*1024):.0f} MB exceeds limit of "
                        f"{MAX_XLSX_UNCOMPRESSED_MB} MB"
                    )
    except ValueError:
        raise
    except zipfile.BadZipFile:
        raise ValueError("XLSX file rejected: not a valid ZIP file")
    except Exception:
        raise ValueError("XLSX file rejected: could not inspect archive")


def _count_xlsx_rows(content: bytes) -> int:
    """Count ``<row `` tags across the workbook's sheets without decompressing
    into openpyxl.

    ``_check_xlsx_zip_bomb`` has already bounded how much this can inflate, so
    reading the sheet XML here is safe. The count is *exact* — an earlier
    version took ``max(row_tags, size / 100)`` as a belt-and-braces estimate,
    but the size term over-estimates by roughly 4x on ordinary exports and
    dominated the max, so a legitimate 25k-row import was rejected as "107,119
    rows". Trust the tag count.
    """
    with zipfile.ZipFile(io.BytesIO(content)) as zf:
        names = [n for n in zf.namelist()
                 if n.startswith("xl/worksheets/") and n.endswith(".xml")]
        if not names:
            names = [n for n in zf.namelist()
                     if "sheet" in n.lower() and n.endswith(".xml")]
        rows = 0
        for name in names:
            rows += zf.read(name).count(b"<row ")
    if rows > MAX_XLSX_ESTIMATED_ROWS:
        raise ValueError(
            f"XLSX file rejected: {rows:,} rows exceeds limit "
            f"of {MAX_XLSX_ESTIMATED_ROWS:,}"
        )
    return rows


def import_xlsx(store, content: bytes, mode: str = "merge", dry_run: bool = False) -> dict:
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required for XLSX import. Install with: pip install openpyxl")
    if mode not in ("merge", "replace"):
        raise ValueError(f"Unknown mode: {mode}")

    _check_xlsx_zip_bomb(content)
    _count_xlsx_rows(content)

    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    created = 0
    updated = 0
    skipped = 0

    # Read and parse the whole sheet before touching the store, for the reason
    # import_table already does: replace mode used to delete every requirement
    # and only then start parsing, so an unreadable sheet — or one whose header
    # row we do not recognise — emptied the project and imported nothing.
    # The row budget is capped by _count_xlsx_rows above, so holding the parsed
    # rows is bounded.
    try:
        parsed_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(v is not None for v in row):
                continue
            row_dict = {headers[i]: str(v) if v is not None else ""
                        for i, v in enumerate(row) if i < len(headers)}
            parsed_rows.append(_row_to_req(row_dict))
    except Exception as exc:
        raise ValueError(f"Could not parse the worksheet (nothing was changed): {exc}") from exc
    finally:
        wb.close()

    row_count = len(parsed_rows)
    row_ids = [r.get("id", "").strip() for r in parsed_rows]

    if mode == "replace":
        _require_id_column(row_ids, row_count)

    if dry_run:
        return _dry_run_summary(store, row_ids, "xlsx", mode, row_count)

    if mode == "replace":
        for req in store.list_requirements():
            store.delete_requirement(req["id"])

    for req_data, rid in zip(parsed_rows, row_ids):
        if not rid:
            skipped += 1
            continue
        if store.get_requirement(rid):
            store.update_requirement(rid, req_data)
            updated += 1
        else:
            store.create_requirement(req_data)
            created += 1

    return {
        "created": created, "updated": updated, "skipped": skipped,
        "traces_added": 0, "verification_cases": 0, "format": "xlsx",
    }
