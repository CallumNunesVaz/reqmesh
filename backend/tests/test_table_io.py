
import pytest

from app.services.table_io import export_table, import_table, _req_to_row, _row_to_req


def test_row_to_req_roundtrip():
    req = {
        "id": "REQ-001", "type": "functional", "name": "Login",
        "description": "The system must authenticate users", "status": "approved",
        "priority": "high", "verification_method": "test", "parent": None,
        "relations": [{"type": "refines", "target": "FEAT-001"}],
        "verification_cases": ["VC-001"], "rationale": "Security requirement",
        "source": "ISO 27001", "allocated_to": "auth-module", "baselines": [],
    }
    row = _req_to_row(req)
    back = _row_to_req(row)
    assert back["id"] == "REQ-001"
    assert back["type"] == "functional"
    assert back["status"] == "approved"
    assert back["priority"] == "high"
    assert len(back["relations"]) == 1
    assert back["relations"][0]["type"] == "refines"
    assert back["relations"][0]["target"] == "FEAT-001"
    assert back["verification_cases"] == ["VC-001"]


def test_export_csv(client, project):
    from app.services.yaml_store import YamlStore
    from app.core.config import settings
    from pathlib import Path

    store = YamlStore(Path(settings.data_root) / project)
    store.create_requirement({"id": "REQ-A", "name": "Alpha", "description": "Do alpha"})
    store.create_requirement({"id": "REQ-B", "name": "Beta", "description": "Do beta"})

    csv_content = export_table(store, "csv")
    lines = csv_content.strip().split("\n")
    assert len(lines) >= 3
    assert lines[0].startswith('"id"')
    assert "Alpha" in csv_content
    assert "Beta" in csv_content


def test_export_tsv(client, project):
    from app.services.yaml_store import YamlStore
    from app.core.config import settings
    from pathlib import Path

    store = YamlStore(Path(settings.data_root) / project)
    store.create_requirement({"id": "REQ-T1", "name": "TSV Test"})

    content = export_table(store, "tsv")
    assert "\t" in content
    assert "REQ-T1" in content


def test_import_csv_merge(client, project):
    from app.services.yaml_store import YamlStore
    from app.core.config import settings
    from pathlib import Path

    store = YamlStore(Path(settings.data_root) / project)
    csv_data = '"id","type","name","description","status","priority","verification_method","parent","relations","verification_cases","rationale","source","allocated_to","baselines"\n"REQ-IMP","functional","Import Test","Do import stuff","proposed","medium","test","","","","","","",""'
    summary = import_table(store, csv_data, fmt="csv", mode="merge")
    assert summary["created"] == 1
    req = store.get_requirement("REQ-IMP")
    assert req is not None
    assert req["name"] == "Import Test"


def test_import_csv_update_existing(client, project):
    from app.services.yaml_store import YamlStore
    from app.core.config import settings
    from pathlib import Path

    store = YamlStore(Path(settings.data_root) / project)
    store.create_requirement({"id": "REQ-EX", "name": "Old Name"})

    csv_data = '"id","type","name","description","status","priority","verification_method","parent","relations","verification_cases","rationale","source","allocated_to","baselines"\n"REQ-EX","functional","New Name","Updated","approved","high","test","","","","","","",""'
    summary = import_table(store, csv_data, fmt="csv", mode="merge")
    assert summary["updated"] == 1
    req = store.get_requirement("REQ-EX")
    assert req["name"] == "New Name"


def test_import_csv_replace(client, project):
    from app.services.yaml_store import YamlStore
    from app.core.config import settings
    from pathlib import Path

    store = YamlStore(Path(settings.data_root) / project)
    store.create_requirement({"id": "REQ-OLD", "name": "Old"})
    assert len(store.list_requirements()) == 1

    csv_data = '"id","type","name","description","status","priority","verification_method","parent","relations","verification_cases","rationale","source","allocated_to","baselines"\n"REQ-NEW","functional","New Only","New desc","proposed","medium","test","","","","","","",""'
    import_table(store, csv_data, fmt="csv", mode="replace")
    reqs = store.list_requirements()
    assert len(reqs) == 1
    assert reqs[0]["id"] == "REQ-NEW"


def test_import_skips_empty_id(client, project):
    from app.services.yaml_store import YamlStore
    from app.core.config import settings
    from pathlib import Path

    store = YamlStore(Path(settings.data_root) / project)
    csv_data = '"id","type","name","description"\n"","functional","No ID","Missing ID"'
    summary = import_table(store, csv_data, fmt="csv", mode="merge")
    assert summary["skipped"] == 1


def test_api_import_csv_endpoint(client, project):
    import io
    csv_content = '"id","type","name","description","status","priority","verification_method","parent","relations","verification_cases","rationale","source","allocated_to","baselines"\n"REQ-API","functional","API Import","Imported via API","proposed","medium","test","","","","","","",""'
    res = client.post(
        f"/api/projects/{project}/import",
        data={"format": "csv", "mode": "merge"},
        files={"file": ("test.csv", io.BytesIO(csv_content.encode("utf-8")), "text/csv")},
    )
    assert res.status_code == 200
    data = res.json()
    assert data["created"] == 1


def test_api_download_csv(client, project):
    from app.services.yaml_store import YamlStore
    from app.core.config import settings
    from pathlib import Path

    store = YamlStore(Path(settings.data_root) / project)
    store.create_requirement({"id": "REQ-DL", "name": "Download Test"})

    res = client.get(f"/api/projects/{project}/publish/download?format=csv")
    assert res.status_code == 200
    content = res.content.decode("utf-8")
    assert "REQ-DL" in content
    assert content.startswith('"id"')


# ── Dry-run previews ─────────────────────────────────────────────────────────
#
# Both table formats must return one shape. They used to disagree: csv reported
# would_create/would_update and zeroed created/updated, while xlsx put the same
# counts in created/updated and omitted the would_* keys entirely — so a caller
# reading one was silently wrong about the other.

HEADER = ('"id","type","name","description","status","priority",'
          '"verification_method","parent","relations","verification_cases",'
          '"rationale","source","allocated_to","baselines"')

DRY_RUN_KEYS = {
    "created", "updated", "skipped", "traces_added", "verification_cases",
    "format", "dry_run", "would_create", "would_update", "would_delete", "rows",
}


def _store(project):
    from app.services.yaml_store import YamlStore
    from app.core.config import settings
    from pathlib import Path

    return YamlStore(Path(settings.data_root) / project)


def _csv(*ids: str) -> str:
    rows = [f'"{rid}","functional","Name {rid}","Desc","proposed","medium","test","","","","","","",""'
            for rid in ids]
    return "\n".join([HEADER, *rows])


def _xlsx(*ids: str) -> bytes:
    import io as _io
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["id", "type", "name", "description"])
    for rid in ids:
        ws.append([rid, "functional", f"Name {rid}", "Desc"])
    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _snapshot(store):
    return sorted((r["id"], r.get("name")) for r in store.list_requirements())


def test_dry_run_merge_reports_creates_and_updates(client, project):
    store = _store(project)
    store.create_requirement({"id": "REQ-EX", "name": "Existing"})

    summary = import_table(store, _csv("REQ-EX", "REQ-NEW"), fmt="csv",
                           mode="merge", dry_run=True)

    assert set(summary) == DRY_RUN_KEYS
    assert summary["dry_run"] is True
    assert summary["created"] == 0 and summary["updated"] == 0
    assert summary["would_create"] == 1
    assert summary["would_update"] == 1
    assert summary["would_delete"] == 0
    assert summary["rows"] == 2


def test_dry_run_replace_reports_creates_as_well_as_deletes(client, project):
    """Replace mode used to return early and report only would_delete, so the
    preview said nothing about what the file would add."""
    store = _store(project)
    store.create_requirement({"id": "REQ-OLD-1", "name": "Old one"})
    store.create_requirement({"id": "REQ-OLD-2", "name": "Old two"})

    summary = import_table(store, _csv("REQ-A", "REQ-B", "REQ-C"), fmt="csv",
                           mode="replace", dry_run=True)

    assert summary["would_delete"] == 2
    # Everything is wiped first, so every id'd row is a create — matching against
    # the current store would be a lie.
    assert summary["would_create"] == 3
    assert summary["would_update"] == 0
    assert summary["rows"] == 3


def test_dry_run_replace_counts_an_overwritten_id_as_a_create(client, project):
    """An id present in both the store and the file is still a create in replace
    mode, because the store copy is deleted before the file is read."""
    store = _store(project)
    store.create_requirement({"id": "REQ-SAME", "name": "Old"})

    summary = import_table(store, _csv("REQ-SAME"), fmt="csv",
                           mode="replace", dry_run=True)

    assert summary["would_create"] == 1
    assert summary["would_update"] == 0
    assert summary["would_delete"] == 1


def test_dry_run_repeated_id_in_file_is_create_then_update(client, project):
    store = _store(project)

    summary = import_table(store, _csv("REQ-DUP", "REQ-DUP"), fmt="csv",
                           mode="merge", dry_run=True)

    assert summary["would_create"] == 1
    assert summary["would_update"] == 1


def test_dry_run_skips_blank_ids_without_counting_them_as_creates(client, project):
    store = _store(project)

    summary = import_table(store, _csv("REQ-OK", ""), fmt="csv",
                           mode="merge", dry_run=True)

    assert summary["skipped"] == 1
    assert summary["would_create"] == 1
    assert summary["rows"] == 2


@pytest.mark.parametrize("mode", ["merge", "replace"])
def test_dry_run_csv_mutates_nothing(client, project, mode):
    store = _store(project)
    store.create_requirement({"id": "REQ-KEEP", "name": "Keep me"})
    before = _snapshot(store)

    import_table(store, _csv("REQ-KEEP", "REQ-NEW"), fmt="csv",
                 mode=mode, dry_run=True)

    assert _snapshot(store) == before


@pytest.mark.parametrize("mode", ["merge", "replace"])
def test_dry_run_xlsx_mutates_nothing(client, project, mode):
    from app.services.table_io import import_xlsx

    store = _store(project)
    store.create_requirement({"id": "REQ-KEEP", "name": "Keep me"})
    before = _snapshot(store)

    import_xlsx(store, _xlsx("REQ-KEEP", "REQ-NEW"), mode=mode, dry_run=True)

    assert _snapshot(store) == before


def test_dry_run_xlsx_uses_the_would_keys_not_created_updated(client, project):
    """The behaviour change: xlsx reported its counts in created/updated."""
    from app.services.table_io import import_xlsx

    store = _store(project)
    store.create_requirement({"id": "REQ-EX", "name": "Existing"})

    summary = import_xlsx(store, _xlsx("REQ-EX", "REQ-NEW"), mode="merge", dry_run=True)

    assert set(summary) == DRY_RUN_KEYS
    assert summary["created"] == 0 and summary["updated"] == 0
    assert summary["would_update"] == 1
    assert summary["would_create"] == 1
    assert summary["format"] == "xlsx"


def test_dry_run_xlsx_rows_excludes_header_and_blank_rows(client, project):
    import io as _io
    import openpyxl
    from app.services.table_io import import_xlsx

    store = _store(project)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["id", "type", "name", "description"])
    ws.append(["REQ-1", "functional", "One", "d"])
    ws.append([None, None, None, None])
    ws.append(["REQ-2", "functional", "Two", "d"])
    buf = _io.BytesIO()
    wb.save(buf)

    summary = import_xlsx(store, buf.getvalue(), mode="merge", dry_run=True)

    assert summary["rows"] == 2
    assert summary["would_create"] == 2


def test_dry_run_replace_refuses_a_file_with_no_id_column(client, project):
    """A preview promising `would_delete: N` for a file the real import rejects
    is worse than no preview — it is what talks the user into committing."""
    store = _store(project)
    store.create_requirement({"id": "REQ-OLD", "name": "Old"})

    bad = '"name","description"\n"No id here","desc"'
    with pytest.raises(ValueError, match="No 'id' column recognised"):
        import_table(store, bad, fmt="csv", mode="replace", dry_run=True)

    assert len(store.list_requirements()) == 1


def test_xlsx_replace_refuses_a_file_with_no_id_column(client, project):
    """xlsx had no such guard at all: replace mode deleted every requirement
    before reading a single data row, so an unrecognised header emptied the
    project and imported nothing."""
    import io as _io
    import openpyxl
    from app.services.table_io import import_xlsx

    store = _store(project)
    store.create_requirement({"id": "REQ-OLD", "name": "Old"})

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["name", "description"])
    ws.append(["No id here", "desc"])
    buf = _io.BytesIO()
    wb.save(buf)

    with pytest.raises(ValueError, match="No 'id' column recognised"):
        import_xlsx(store, buf.getvalue(), mode="replace", dry_run=False)

    assert len(store.list_requirements()) == 1


# ── Import history ────────────────────────────────────────────────────────────

def _csv_body(*ids: str) -> str:
    rows = [f'"{rid}","functional","Name {rid}","Desc","proposed","medium","test","","","","","","",""'
            for rid in ids]
    return "\n".join([HEADER, *rows])


def _history_ids(store, item_id: str) -> list[str]:
    entries = store.list_history(item_id)
    return [e["action"] for e in entries]


def test_import_merge_writes_create_history(client, project):
    store = _store(project)

    import_table(store, _csv_body("REQ-H1"), fmt="csv", mode="merge", username="importer")

    assert _history_ids(store, "REQ-H1") == ["create"]


def test_import_merge_writes_update_history(client, project):
    store = _store(project)
    store.create_requirement({"id": "REQ-H2", "name": "Old Name"})

    import_table(store, _csv_body("REQ-H2"), fmt="csv", mode="merge", username="importer")

    actions = _history_ids(store, "REQ-H2")
    assert "update" in actions


def test_import_replace_writes_delete_history(client, project):
    store = _store(project)
    store.create_requirement({"id": "REQ-OLD", "name": "Old"})

    import_table(store, _csv_body("REQ-NEW"), fmt="csv", mode="replace", username="importer")

    actions = _history_ids(store, "REQ-OLD")
    assert actions == ["delete"]


def test_import_dry_run_writes_no_history(client, project):
    store = _store(project)
    store.create_requirement({"id": "REQ-DRY", "name": "Dry"})

    import_table(store, _csv_body("REQ-DRY", "REQ-NEW"), fmt="csv", mode="merge",
                 dry_run=True, username="importer")

    # History should be empty — dry run writes nothing.
    assert _history_ids(store, "REQ-DRY") == []


def test_reimport_identical_rows_adds_no_new_update_entries(client, project):
    store = _store(project)

    import_table(store, _csv_body("REQ-SAME"), fmt="csv", mode="merge", username="importer")
    first_actions = len(store.list_history("REQ-SAME"))

    import_table(store, _csv_body("REQ-SAME"), fmt="csv", mode="merge", username="importer")
    second_actions = len(store.list_history("REQ-SAME"))

    assert second_actions == first_actions
